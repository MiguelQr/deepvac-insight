"""Tests for data_service.make_report_xlsx() -- the only report format this
app currently generates. (The old PDF export mentioned in project history
was removed entirely; there's no HTML export code path either, just this
single-sheet .xlsx via pandas + openpyxl.)"""

import openpyxl
import pytest

from app.services import data_service

pytestmark = pytest.mark.integration


@pytest.fixture
def fake_workspace(tmp_path, monkeypatch):
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    monkeypatch.setenv("DEEPVAC_DATA_ROOT", str(workspace))
    return workspace


def _write_run(root, name, rows=5):
    run_dir = root / name
    run_dir.mkdir(parents=True, exist_ok=True)
    lines = ["run_id,timestamp,temp,temp_ref"]
    for i in range(rows):
        lines.append(f"{name},{1700000000 + i * 2},{20.0 + i},60.0")
    (run_dir / "run_samples.csv").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return run_dir


def test_make_report_xlsx_creates_nonempty_file(deepvac_data_dir, fake_workspace, tmp_path):
    _write_run(fake_workspace, "run_a", rows=5)
    data_service.sync_cache()
    key = data_service.load_cached_runs()[0]["key"]

    output_path = tmp_path / "report.xlsx"
    data_service.make_report_xlsx(key, output_path)

    assert output_path.exists()
    assert output_path.stat().st_size > 0


def test_make_report_xlsx_sheet_name_is_raw_data(deepvac_data_dir, fake_workspace, tmp_path):
    _write_run(fake_workspace, "run_a", rows=3)
    data_service.sync_cache()
    key = data_service.load_cached_runs()[0]["key"]

    output_path = tmp_path / "report.xlsx"
    data_service.make_report_xlsx(key, output_path)

    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames == ["Raw Data"]


def test_make_report_xlsx_headers_and_row_count(deepvac_data_dir, fake_workspace, tmp_path):
    _write_run(fake_workspace, "run_a", rows=4)
    data_service.sync_cache()
    key = data_service.load_cached_runs()[0]["key"]

    output_path = tmp_path / "report.xlsx"
    data_service.make_report_xlsx(key, output_path)

    wb = openpyxl.load_workbook(output_path)
    sheet = wb["Raw Data"]
    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header_row == ["run_id", "timestamp", "temp", "temp_ref"]
    assert sheet.max_row == 5  # header + 4 data rows


def test_make_report_xlsx_representative_values(deepvac_data_dir, fake_workspace, tmp_path):
    # Characterizes current (possibly worth revisiting, not changed here)
    # behavior: run_table()'s rows come straight from csv.DictReader with
    # no numeric conversion, and make_report_xlsx() doesn't convert them
    # either -- so numeric-looking columns land in the .xlsx as *text*
    # ("20.0"), not real Excel numbers. That means a generated report isn't
    # directly usable for in-Excel sorting/formulas on those columns without
    # the user first converting them -- worth flagging separately, since
    # deciding whether "Raw Data" should mean "typed" or "verbatim from CSV"
    # is a product decision, not something to change silently in a test pass.
    _write_run(fake_workspace, "run_a", rows=2)
    data_service.sync_cache()
    key = data_service.load_cached_runs()[0]["key"]

    output_path = tmp_path / "report.xlsx"
    data_service.make_report_xlsx(key, output_path)

    wb = openpyxl.load_workbook(output_path)
    sheet = wb["Raw Data"]
    first_data_row = [cell.value for cell in list(sheet.iter_rows(min_row=2, max_row=2))[0]]
    assert first_data_row[0] == "run_a"
    assert first_data_row[2] == "20.0"  # temp -- stored as text, not a number
